"""
Template Python Provider for the Rust "provider" server.

Drop this file somewhere under your project base dir, e.g.:
  <base>/providers/my_plugins/template_provider.py

Load it at runtime via HTTP:
  POST /plugins/load
  {
    "module": "providers.my_plugins.template_provider",
    "class": "Provider",
    "name": "template",
    "project_base_dir": "<base>"
  }

Or from the client shell:
  :http http://127.0.0.1:7070
  :loadpy module=providers.my_plugins.template_provider class=Provider base=<base> name=template

Contract
--------
Your Provider class must expose:
  - name(self) -> str
  - fetch_entities(self, request_json: dict) -> list[EntityDict]
  - (optional) stitch(self, filters_json: list[dict]) -> EntityDict

Where EntityDict matches the Rust struct fields (all strings):
    {
      "id": "source:key...",
      "source": "your_source_name",
      "tags": "[\"k=v\", ...]",        # stringified JSON array of tags
      "data": "[{...}, {...}]",          # stringified JSON array of records
      "etag": "<hex>",
      "fetched_at": "RFC3339",
      "refresh_after": "RFC3339",
      "state": "ready",
      "last_error": "",
      "updated_at": "RFC3339"
    }

`request_json` mirrors the Rust-side EntityInProvider variants, for example:
  { "GetEntity": { "id": "..." } }
  { "GetEntities": { "ids": ["...", ...] } }
  { "GetAllEntities": {} }
  { "SearchEntities": { "query": [ {"Ticker": "AAPL"},
                                   {"DateRange": {"start":"2025-09-01T00:00:00Z",
                                                  "end":"2025-10-01T00:00:00Z"}} ] } }

You are free to interpret filters specific to your provider, e.g. { "CsvPath": "/path/file.csv" }.
"""

from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple


# ====== Small helpers you can reuse/modify ======

RFC3339 = "%Y-%m-%dT%H:%M:%SZ"  # naive Zulu; adjust if you prefer full offset format


def now_rfc3339(dt: Optional[datetime] = None) -> str:
    dt = dt or datetime.now(timezone.utc)
    # normalize to '...Z' (no offset) for consistency with Rust example
    return dt.astimezone(timezone.utc).strftime(RFC3339)


def make_etag(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8")).hexdigest()


def tags_to_string(tags: Dict[str, str]) -> str:
    # Server expects a JSON-encoded array of "k=v" strings
    arr = [f"{k}={v}" for k, v in tags.items()]
    return json.dumps(arr, separators=(",", ":"))


def records_to_string(records: List[Dict[str, Any]]) -> str:
    return json.dumps(records, separators=(",", ":"))


def make_entity(
    *,
    source: str,
    entity_id: str,
    tags: Dict[str, str],
    records: List[Dict[str, Any]],
    ttl_days: int = 1,
) -> Dict[str, str]:
    """Build a single Entity dict with required fields as strings."""
    data_str = records_to_string(records)
    etag = make_etag(data_str)
    fetched = now_rfc3339()
    updated = fetched
    refresh_after = now_rfc3339(datetime.now(timezone.utc) + timedelta(days=ttl_days))

    return {
        "id": entity_id,
        "source": source,
        "tags": tags_to_string(tags),
        "data": data_str,
        "etag": etag,
        "fetched_at": fetched,
        "refresh_after": refresh_after,
        "state": "ready",
        "last_error": "",
        "updated_at": updated,
    }


# ====== Optional: CSV / Excel ingestion helpers (bi-directional) ======

def load_records_from_csv(path: str, max_rows: Optional[int] = None) -> List[Dict[str, Any]]:
    """
    Lightweight CSV reader without extra dependencies.
    If you prefer pandas, replace with pd.read_csv(...).to_dict(orient="records").
    """
    import csv
    rows: List[Dict[str, Any]] = []
    with open(path, "r", newline="", encoding="utf-8") as f:
        rdr = csv.DictReader(f)
        for i, row in enumerate(rdr):
            rows.append(dict(row))
            if max_rows is not None and i + 1 >= max_rows:
                break
    return rows


def load_records_from_excel(path: str, sheet: Optional[str] = None, max_rows: Optional[int] = None) -> List[Dict[str, Any]]:
    """
    Excel reader via openpyxl (install: pip install openpyxl).
    """
    from openpyxl import load_workbook  # type: ignore
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows: List[Dict[str, Any]] = []

    # Ensure the worksheet is valid
    if ws is None:
        raise ValueError(f"Worksheet '{sheet}' not found in the workbook.")

    # first row -> headers
    it = ws.iter_rows(values_only=True)
    headers = next(it, None)
    if not headers:
        return rows
    headers = [str(h) if h is not None else "" for h in headers]

    for i, r in enumerate(it):
        obj = {}
        for k, v in zip(headers, r):
            obj[str(k)] = v
        rows.append(obj)
        if max_rows is not None and i + 1 >= max_rows:
            break
    return rows


# ====== Provider implementation template ======

@dataclass
class DateRange:
    start: str
    end: str


class Provider:
    """
    Minimal provider template.

    Implement your data fetch in `_search_impl()` and optional `_stitch_impl()`.
    Optionally accept CSV/XLSX paths via filters in request_json.
    """

    def __init__(self) -> None:
        # Initialize any clients, tokens, etc.
        self._source_name = "template_provider"

    # --- Required ---
    def name(self) -> str:
        """Return the provider's canonical name (used as default alias when registering)."""
        return self._source_name

    # --- Required ---
    def fetch_entities(self, request_json: Dict[str, Any]) -> List[Dict[str, str]]:
        """
        Interpret the request and return a list of Entity dicts.
        You may support any subset of the Rust-side variants that make sense.
        """
        # Detect variant by top-level key
        if "GetEntity" in request_json:
            # Example: return a single entity by id sourced from local CSV for demo
            entity_id = request_json["GetEntity"].get("id", "template:demo")
            records = [
                {"timestamp": "2025-09-01T00:00:00Z", "value": 1},
                {"timestamp": "2025-09-02T00:00:00Z", "value": 2},
            ]
            ent = make_entity(
                source=self._source_name,
                entity_id=entity_id,
                tags={"demo": "true"},
                records=records,
            )
            return [ent]

        if "GetAllEntities" in request_json:
            # Return an empty list by default
            return []

        if "GetEntities" in request_json:
            # Typically you'd look up by ID(s) from your own storage. As a template, return empty.
            return []

        if "SearchEntities" in request_json:
            # Interpret "query" field with custom filters.
            filters = request_json["SearchEntities"].get("query", [])
            # Example: support either a CSV path or DateRange+Ticker, etc.
            csv_path = _find_filter(filters, "CsvPath")
            xlsx_path = _find_filter(filters, "ExcelPath")

            if csv_path:
                records = load_records_from_csv(csv_path)
                ent = make_entity(
                    source=self._source_name,
                    entity_id=f"{self._source_name}:csv:{csv_path}",
                    tags={"csv": csv_path},
                    records=records,
                )
                return [ent]

            if xlsx_path:
                records = load_records_from_excel(xlsx_path)
                ent = make_entity(
                    source=self._source_name,
                    entity_id=f"{self._source_name}:xlsx:{xlsx_path}",
                    tags={"xlsx": xlsx_path},
                    records=records,
                )
                return [ent]

            # Example: Ticker + DateRange
            ticker = _find_filter(filters, "Ticker")
            dr = _find_date_range(filters)

            # Produce synthetic rows for demo
            records = _synth_timeseries(ticker or "DEMO", dr)
            ent = make_entity(
                source=self._source_name,
                entity_id=f"{self._source_name}:{ticker or 'DEMO'}:{(dr.start if dr else 'start')}..{(dr.end if dr else 'end')}",
                tags={
                    "ticker": ticker or "DEMO",
                    "from": (dr.start if dr else ""),
                    "to": (dr.end if dr else ""),
                },
                records=records,
            )
            return [ent]

        # Unknown request
        return []

    # --- Optional ---
    def stitch(self, filters_json: List[Dict[str, Any]]) -> Dict[str, str]:
        """
        Optional: Given a list of filters, return a single stitched Entity.
        If not needed, you can remove this method; the Rust adapter will report "not supported".
        """
        # Example behavior: accept multiple CsvPath filters and concat rows
        csv_paths = _find_all(filters_json, "CsvPath")
        all_records: List[Dict[str, Any]] = []
        for p in csv_paths:
            try:
                all_records.extend(load_records_from_csv(p))
            except Exception as e:  # keep template simple
                all_records.append({"path": p, "error": str(e)})

        ent = make_entity(
            source=self._source_name,
            entity_id=f"{self._source_name}:stitch:csv",
            tags={"kind": "stitch", "inputs": ";".join(csv_paths)},
            records=all_records,
        )
        return ent


# ====== Simple filter helpers ======

def _find_filter(filters: List[Dict[str, Any]], key: str) -> Optional[str]:
    """
    Find a filter like { key: "value" } and return the value as str if present.
    """
    for f in filters:
        if key in f:
            v = f[key]
            if isinstance(v, str):
                return v
    return None


def _find_all(filters: List[Dict[str, Any]], key: str) -> List[str]:
    out: List[str] = []
    for f in filters:
        if key in f and isinstance(f[key], str):
            out.append(f[key])
    return out


def _find_date_range(filters: List[Dict[str, Any]]) -> Optional[DateRange]:
    for f in filters:
        if "DateRange" in f and isinstance(f["DateRange"], dict):
            dr = f["DateRange"]
            start = str(dr.get("start", ""))
            end = str(dr.get("end", ""))
            return DateRange(start=start, end=end)
    return None


def _synth_timeseries(ticker: str, dr: Optional[DateRange]) -> List[Dict[str, Any]]:
    # Construct a tiny timeseries for demonstration if no CSV/XLSX provided.
    # If no DateRange, produce 3 points.
    if not dr or not dr.start or not dr.end:
        return [
            {"timestamp": "2025-01-01T00:00:00Z", "ticker": ticker, "value": 1},
            {"timestamp": "2025-01-02T00:00:00Z", "ticker": ticker, "value": 2},
            {"timestamp": "2025-01-03T00:00:00Z", "ticker": ticker, "value": 3},
        ]
    # Otherwise just echo the boundaries as example
    return [
        {"timestamp": dr.start, "ticker": ticker, "value": 1},
        {"timestamp": dr.end, "ticker": ticker, "value": 2},
    ]
